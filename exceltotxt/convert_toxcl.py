import openpyxl
import os
import sys

#Read text values from file
with open (os.path.join(sys.path[0], 'data.txt'), 'r') as file:
    values = file.read().splitlines()

#Create new Excel workbook
workbook = openpyxl.Workbook()
worksheet = workbook.active

#write values to Excel worksheet
for i, value in enumerate(values, start=1):
    worksheet.cell(row=i, column=1, value=value)


#Save Excel workbook to file
workbook.save('data.xlsx')

#Read values from Excel worksheet and print to console
for row in worksheet.iter_rows(min_row=1, max_col=3):
    for cell in row:
        print(cell.value)